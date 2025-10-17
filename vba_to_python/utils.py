"""
Utility helpers shared by the action modules.
"""

from __future__ import annotations

import os
import re
from contextlib import contextmanager
from pathlib import Path
from typing import Iterable, Iterator, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .constants import (
    DATA_UPDATE_MAX_COLUMNS,
    DATA_UPDATE_START_ROW,
    EXCEL_EXTENSIONS,
    FILE_LIST_COLUMN,
    FILE_LIST_START_ROW,
    IOCHANGE_MAX_COL,
    IOCHANGE_MAX_ROW,
    IOCHANGE_MIN_COL,
    IOCHANGE_MIN_ROW,
    SHEET_DATA_UPDATE,
    SHEET_FILES,
    SHEET_IO_NAME,
)


def normalize_path(value: Optional[str]) -> Optional[str]:
    """Strip whitespace and convert empty strings to ``None``."""
    if value is None:
        return None
    normalized = str(value).strip()
    return normalized or None


def ensure_trailing_sep(path_str: str) -> str:
    """Guarantee that a filesystem path ends with the OS separator."""
    if path_str.endswith(("\\", "/")):
        return path_str
    return path_str + os.sep


def list_excel_filenames(folder: Path) -> List[str]:
    """Return sorted Excel filenames residing in ``folder``."""
    files: List[str] = []
    for entry in folder.iterdir():
        if not entry.is_file():
            continue
        if entry.suffix.lower() in EXCEL_EXTENSIONS:
            files.append(entry.name)
    return sorted(files)


def increment_suffix(value: str) -> str:
    """Increment the numeric suffix after the last underscore."""
    match = re.search(r"_(\d+)$", value)
    if not match:
        raise ValueError(f"잘못된 ID 형식입니다: {value}")
    number = int(match.group(1)) + 1
    return f"{value[:match.start(1)]}{number:02d}"


def is_valid_cell_address(address: str) -> bool:
    """Return True if ``address`` looks like an Excel cell reference."""
    return bool(re.fullmatch(r"[A-Za-z]{1,3}[1-9][0-9]{0,6}", address.strip()))


def load_control_workbook(path: Path) -> Workbook:
    """Load the control workbook (contains the `파일` sheet)."""
    if not path.exists():
        raise FileNotFoundError(f"컨트롤 통합문서를 찾을 수 없습니다: {path}")
    keep_vba = path.suffix.lower() in {".xlsm", ".xlam"}
    return load_workbook(path, keep_vba=keep_vba)


def load_target_workbook(path: Path) -> Workbook:
    """Load a target workbook that will be modified."""
    if not path.exists():
        raise FileNotFoundError(f"파일을 찾을 수 없습니다: {path}")
    suffix = path.suffix.lower()
    if suffix in {".xlsx"}:
        return load_workbook(path)
    if suffix in {".xlsm", ".xlam"}:
        return load_workbook(path, keep_vba=True)
    raise ValueError(
        f"지원하지 않는 파일 형식입니다 (xlsx/xlsm만 지원): {path.name}"
    )


@contextmanager
def open_control_workbook(path: Path) -> Iterator[Workbook]:
    wb = load_control_workbook(path)
    try:
        yield wb
    finally:
        wb.close()


def get_required_sheet(wb: Workbook, sheet_name: str) -> Worksheet:
    if sheet_name not in wb.sheetnames:
        available = ", ".join(wb.sheetnames)
        raise ValueError(f"'{sheet_name}' 시트를 찾을 수 없습니다. (현재 시트: {available})")
    return wb[sheet_name]


def clear_column(ws: Worksheet, column_index: int, start_row: int) -> None:
    """Clear values from ``start_row`` downwards within ``column_index``."""
    max_row = ws.max_row
    if max_row < start_row:
        return
    for row in range(start_row, max_row + 1):
        ws.cell(row=row, column=column_index).value = None


def extract_file_list(ws_files: Worksheet) -> List[str]:
    """Read the file list from the control sheet."""
    files: List[str] = []
    row = FILE_LIST_START_ROW
    while True:
        cell_value = normalize_path(ws_files.cell(row=row, column=FILE_LIST_COLUMN).value)
        if not cell_value:
            break
        files.append(cell_value)
        row += 1
    return files


def read_replacements(ws_map: Worksheet) -> List[Tuple[str, str]]:
    """Read the IO replacement mapping from the IO_name sheet."""
    mappings: List[Tuple[str, str]] = []
    for row in ws_map.iter_rows(min_row=1, max_col=2):
        source = normalize_path(row[0].value)
        target = row[1].value if len(row) > 1 else ""
        if source:
            mappings.append((source, "" if target is None else str(target)))
    return mappings


def clear_data_update(ws_data_update: Worksheet) -> None:
    """Clear data rows on the data_update sheet (rows 2 and below)."""
    max_row = ws_data_update.max_row
    if max_row < DATA_UPDATE_START_ROW:
        return
    for row in range(DATA_UPDATE_START_ROW, max_row + 1):
        for col in range(1, DATA_UPDATE_MAX_COLUMNS + 1):
            ws_data_update.cell(row=row, column=col).value = None


def iter_iochange_cells(ws: Worksheet):
    """
    Iterate over the A5:M700 range cells (inclusive) used by the IO change macro.
    """
    return ws.iter_rows(
        min_row=IOCHANGE_MIN_ROW,
        max_row=IOCHANGE_MAX_ROW,
        min_col=IOCHANGE_MIN_COL,
        max_col=IOCHANGE_MAX_COL,
    )


def ensure_directory(path: Path) -> None:
    """Create the directory if it does not exist."""
    path.mkdir(parents=True, exist_ok=True)

