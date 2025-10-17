#!/usr/bin/env python3
"""
SMB Precheck Automation (Python Edition)
=======================================

SMB 결재 서류 사전 검토 프로세스를 Excel 통합문서와 연계해 자동화합니다.
기존 VBA 매크로와 동일한 작업 순서를 CLI에서 제공하며, 작업 대상은 `_processed`
사본에만 적용되고 원본은 `_backup` 폴더로 보존됩니다.

사용 예:
    python tools/smb_precheck.py scan --control control.xlsx --base C:\Data
    python tools/smb_precheck.py update-ids --control control.xlsx
    python tools/smb_precheck.py io-change --control control.xlsx
    python tools/smb_precheck.py value-find --control control.xlsx
    python tools/smb_precheck.py change-value --control control.xlsx
    python tools/smb_precheck.py list-subfolders --control control.xlsx
"""

from __future__ import annotations

import argparse
import os
import re
import shutil
import sys
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import List, Optional, Sequence, Tuple

try:
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill
except ImportError as exc:  # pragma: no cover
    print("openpyxl 패키지가 설치되어 있어야 합니다. `pip install openpyxl` 후 다시 실행하세요.", file=sys.stderr)
    raise


SHEET_FILES = "파일"
SHEET_IO_NAME = "IO_name"
SHEET_DATA_UPDATE = "data_update"
SHEET_SCRIPT_MOVE = "script_move"

BASE_PATH_CELL = "B2"
INCLUDE_FILTER_CELL = "B4"
EXCLUDE_FILTER_CELL = "B5"
FIND_VALUE_CELL = "B10"
TARGET_SHEET_CELL = "B12"

FILE_TABLE_HEADER_ROW = 7
FILE_TABLE_START_ROW = 8

COL_FILE_NAME = 1
COL_ORIGINAL_PATH = 2
COL_INCLUDE = 3
COL_STATUS = 4
COL_MESSAGE = 5

SUCCESS_FILL = PatternFill(fill_type="solid", fgColor="C6EFCE")
FAIL_FILL = PatternFill(fill_type="solid", fgColor="FFC7CE")
CLEAR_FILL = PatternFill(fill_type=None)

DEFAULT_LOG_DIR = Path("vba") / "log"


@dataclass
class SelectedFile:
    row: int
    file_name: str
    original_path: Path
    processed_path: Optional[Path] = None


def ensure_control_workbook(control_path: Path):
    if not control_path.exists():
        raise FileNotFoundError(f"관리 통합문서를 찾을 수 없습니다: {control_path}")
    wb = load_workbook(control_path)
    for sheet in (SHEET_FILES, SHEET_IO_NAME, SHEET_DATA_UPDATE, SHEET_SCRIPT_MOVE):
        if sheet not in wb.sheetnames:
            wb.close()
            raise ValueError(f"통합문서에 '{sheet}' 시트가 없습니다.")
    return wb


def normalize_path(value: Optional[str]) -> Optional[str]:
    if value is None:
        return None
    value = str(value).strip()
    return value or None


def ensure_trailing_slash(path_str: str) -> str:
    if path_str.endswith(("\\", "/")):
        return path_str
    return path_str + os.sep


def clear_table(ws, start_row: int, start_col: int, end_col: int):
    max_row = ws.max_row
    if max_row < start_row:
        return
    for row in ws.iter_rows(min_row=start_row, max_row=max_row, min_col=start_col, max_col=end_col):
        for cell in row:
            cell.value = None
            cell.fill = CLEAR_FILL


def parse_filter_tokens(raw: Optional[str]) -> List[str]:
    if raw is None:
        return []
    tokens = [token.strip().lower() for token in raw.split(";")]
    return [token for token in tokens if token]


def matches_tokens(name: str, includes: Sequence[str], excludes: Sequence[str]) -> bool:
    lowered = name.lower()
    if includes and not any(token in lowered for token in includes):
        return False
    if excludes and any(token in lowered for token in excludes):
        return False
    return True


def read_bool(cell_value) -> Optional[bool]:
    if isinstance(cell_value, bool):
        return cell_value
    if cell_value is None:
        return None
    text = str(cell_value).strip().lower()
    if text in {"true", "y", "yes", "1"}:
        return True
    if text in {"false", "n", "no", "0"}:
        return False
    return None


def prepare_processed_copy(original_path: Path) -> Path:
    if not original_path.exists():
        raise FileNotFoundError(f"파일을 찾을 수 없습니다: {original_path}")
    base_dir = original_path.parent
    backup_dir = base_dir / "_backup"
    processed_dir = base_dir / "_processed"
    backup_dir.mkdir(exist_ok=True)
    processed_dir.mkdir(exist_ok=True)

    backup_path = backup_dir / original_path.name
    processed_path = processed_dir / original_path.name
    if not backup_path.exists():
        shutil.copy2(original_path, backup_path)
    shutil.copy2(original_path, processed_path)
    return processed_path


def load_target_workbook(path: Path):
    keep_vba = path.suffix.lower() == ".xlsm"
    return load_workbook(path, keep_vba=keep_vba)


def ensure_log_dir(base: Path = DEFAULT_LOG_DIR) -> Path:
    base = base.resolve()
    base.mkdir(parents=True, exist_ok=True)
    return base


def write_log_if_needed(messages: List[str], base_dir: Path = DEFAULT_LOG_DIR) -> Optional[Path]:
    if not messages:
        return None
    log_dir = ensure_log_dir(base_dir)
    log_path = log_dir / f"SMB_{datetime.now():%Y%m%d_%H%M%S}.log"
    log_path.write_text(
        "SMB Precheck Log\n"
        f"Created: {datetime.now():%Y-%m-%d %H:%M:%S}\n"
        + "-" * 60 + "\n"
        + "\n".join(messages),
        encoding="utf-8",
    )
    return log_path


def list_excel_files(control_path: Path, base_folder: Optional[Path]) -> int:
    wb = ensure_control_workbook(control_path)
    ws = wb[SHEET_FILES]

    if base_folder:
        ws[BASE_PATH_CELL].value = ensure_trailing_slash(str(base_folder.resolve()))

    base_value = normalize_path(ws[BASE_PATH_CELL].value)
    if not base_value:
        wb.close()
        raise ValueError("파일 시트의 B2에 기본 경로가 비어 있습니다.")

    base_dir = Path(base_value)
    if not base_dir.exists():
        wb.close()
        raise FileNotFoundError(f"기본 경로를 찾을 수 없습니다: {base_dir}")

    include_tokens = parse_filter_tokens(normalize_path(ws[INCLUDE_FILTER_CELL].value))
    exclude_tokens = parse_filter_tokens(normalize_path(ws[EXCLUDE_FILTER_CELL].value))

    clear_table(ws, FILE_TABLE_START_ROW, COL_FILE_NAME, COL_MESSAGE)

    count = 0
    row = FILE_TABLE_START_ROW
    for file_path in sorted(base_dir.glob("*.xls*")):
        ws.cell(row=row, column=COL_FILE_NAME).value = file_path.name
        ws.cell(row=row, column=COL_ORIGINAL_PATH).value = str(file_path.resolve())
        ws.cell(row=row, column=COL_INCLUDE).value = matches_tokens(file_path.name, include_tokens, exclude_tokens)
        ws.cell(row=row, column=COL_STATUS).fill = CLEAR_FILL
        ws.cell(row=row, column=COL_STATUS).value = None
        ws.cell(row=row, column=COL_MESSAGE).value = None
        row += 1
        count += 1

    wb.save(control_path)
    wb.close()
    return count


def collect_selected_files(wb, base_dir: Path) -> List[SelectedFile]:
    ws = wb[SHEET_FILES]
    include_tokens = parse_filter_tokens(normalize_path(ws[INCLUDE_FILTER_CELL].value))
    exclude_tokens = parse_filter_tokens(normalize_path(ws[EXCLUDE_FILTER_CELL].value))

    selected: List[SelectedFile] = []
    row = FILE_TABLE_START_ROW
    while True:
        file_name = normalize_path(ws.cell(row=row, column=COL_FILE_NAME).value)
        if not file_name:
            break
        original_str = normalize_path(ws.cell(row=row, column=COL_ORIGINAL_PATH).value)
        original_path = Path(original_str) if original_str else base_dir / file_name

        manual = read_bool(ws.cell(row=row, column=COL_INCLUDE).value)
        if manual is None:
            auto = matches_tokens(file_name, include_tokens, exclude_tokens)
            ws.cell(row=row, column=COL_INCLUDE).value = auto
            include = auto
        else:
            include = manual

        if include:
            selected.append(SelectedFile(row=row, file_name=file_name, original_path=original_path))

        row += 1

    return selected


def clear_status_columns(ws):
    for row in ws.iter_rows(min_row=FILE_TABLE_START_ROW, max_row=ws.max_row,
                            min_col=COL_STATUS, max_col=COL_MESSAGE):
        status_cell, message_cell = row
        status_cell.value = None
        status_cell.fill = CLEAR_FILL
        message_cell.value = None


def mark_status(ws, row: int, success: bool, message: str):
    cell_status = ws.cell(row=row, column=COL_STATUS)
    cell_message = ws.cell(row=row, column=COL_MESSAGE)
    if success:
        cell_status.value = "Success"
        cell_status.fill = SUCCESS_FILL
    else:
        cell_status.value = "Fail"
        cell_status.fill = FAIL_FILL
    cell_message.value = message


def get_control_base_dir(ws_files) -> Path:
    base_value = normalize_path(ws_files[BASE_PATH_CELL].value)
    if not base_value:
        raise ValueError("파일 시트의 B2에 기본 경로가 비어 있습니다.")
    base_dir = Path(base_value)
    if not base_dir.exists():
        raise FileNotFoundError(f"기본 경로를 찾을 수 없습니다: {base_dir}")
    return base_dir


def task_update_ids(target_path: Path) -> str:
    wb = load_target_workbook(target_path)
    if "Test Case" not in wb.sheetnames:
        wb.close()
        raise ValueError("Test Case 시트를 찾을 수 없습니다.")
    ws = wb["Test Case"]

    main_name = normalize_path(ws["A2"].value)
    if not main_name:
        wb.close()
        raise ValueError("A2 셀(메인 식별자)이 비어 있습니다.")

    depth1 = f"{main_name}_00"
    depth2 = f"{depth1}_01"
    ws["A3"].value = depth1
    ws["A4"].value = depth2

    row = 5
    while True:
        cell = ws[f"A{row}"]
        value = normalize_path(cell.value)
        if not value:
            break
        if len(value) == len(depth1):
            depth1 = increment_suffix(depth1)
            cell.value = depth1
            depth2 = f"{depth1}_00"
        elif len(value) == len(depth2):
            descriptor = normalize_path(ws[f"B{row}"].value) or ""
            if "precondition" in descriptor.lower():
                cell.value = depth2
            else:
                depth2 = increment_suffix(depth2)
                cell.value = depth2
        else:
            wb.close()
            raise ValueError(f"예상치 못한 ID 패턴 (행 {row}): {value}")
        row += 1

    wb.save(target_path)
    wb.close()
    return "케이스 ID 재정렬 완료"


def increment_suffix(value: str) -> str:
    match = re.search(r"_(\d+)$", value)
    if not match:
        raise ValueError(f"ID 형식이 올바르지 않습니다: {value}")
    number = int(match.group(1)) + 1
    return f"{value[:match.start(1)]}{number:02d}"


def task_io_change(target_path: Path, replacements: List[Tuple[str, str]]) -> str:
    wb = load_target_workbook(target_path)
    if "Test Case" not in wb.sheetnames:
        wb.close()
        raise ValueError("Test Case 시트를 찾을 수 없습니다.")
    ws = wb["Test Case"]

    for row in ws.iter_rows(min_row=5, max_row=700, min_col=1, max_col=13):
        for cell in row:
            if not isinstance(cell.value, str):
                continue
            text = cell.value
            for before, after in replacements:
                if before and before in text:
                    text = text.replace(before, after)
            cell.value = text

    wb.save(target_path)
    wb.close()
    return "IO 텍스트 치환 완료"


def task_value_find(control_wb, target_path: Path, target_sheet_name: str,
                    find_text: str, data_update_ws) -> str:
    wb = load_target_workbook(target_path)
    if "Test Case" not in wb.sheetnames:
        wb.close()
        raise ValueError("Test Case 시트를 찾을 수 없습니다.")
    ws = wb["Test Case"]

    find_lower = find_text.lower()
    matches = 0
    next_row = data_update_ws.max_row + 1
    if next_row < 2:
        next_row = 2

    for cell in ws.iter_rows(min_col=3, max_col=6, values_only=False):
        for item in cell:
            value = item.value
            if not isinstance(value, str):
                continue
            if find_lower in value.lower():
                matches += 1
                data_update_ws.cell(row=next_row, column=1).value = str(target_path)
                data_update_ws.cell(row=next_row, column=2).value = value
                data_update_ws.cell(row=next_row, column=3).value = target_sheet_name
                right_cell = ws.cell(row=item.row, column=item.column + 1)
                data_update_ws.cell(row=next_row, column=4).value = right_cell.value
                data_update_ws.cell(row=next_row, column=5).value = right_cell.coordinate
                next_row += 1

    wb.close()
    return "일치 {}건".format(matches) if matches else "일치 항목 없음"


def task_change_value(control_wb, log_messages: List[str]) -> Tuple[int, int]:
    ws = control_wb[SHEET_DATA_UPDATE]
    last_row = ws.max_row
    if last_row < 2:
        print("data_update 시트에 처리할 항목이 없습니다.")
        return (0, 0)

    success = 0
    failure = 0
    start = time.time()

    for idx in range(2, last_row + 1):
        file_path = normalize_path(ws.cell(row=idx, column=1).value)
        target_sheet = normalize_path(ws.cell(row=idx, column=3).value)
        cell_addr = normalize_path(ws.cell(row=idx, column=5).value)
        new_value = ws.cell(row=idx, column=6).value

        if not file_path or not target_sheet or not cell_addr:
            continue
        file_path = Path(file_path)
        if not file_path.exists():
            failure += 1
            log_messages.append(f"{file_path} | 파일을 찾을 수 없습니다.")
            continue
        if not is_valid_cell_address(cell_addr):
            failure += 1
            log_messages.append(f"{file_path} | 잘못된 셀 주소: {cell_addr}")
            continue

        try:
            wb = load_target_workbook(file_path)
            if target_sheet not in wb.sheetnames:
                failure += 1
                log_messages.append(f"{file_path} | 대상 시트를 찾을 수 없습니다: {target_sheet}")
                wb.close()
                continue
            wb[target_sheet][cell_addr].value = new_value
            wb.save(file_path)
            wb.close()
            success += 1
        except Exception as exc:  # pragma: no cover
            failure += 1
            log_messages.append(f"{file_path} | {exc}")

        processed = success + failure
        elapsed = time.time() - start
        avg = elapsed / processed if processed else 0
        remaining = avg * (last_row - 1 - processed)
        eta = f"{remaining/60:.1f} min" if remaining >= 60 else f"{remaining:.0f} sec"
        print(f"[Change] {processed}/{last_row-1} 완료 (ETA {eta})", end="\r", flush=True)

    print()
    return success, failure


def list_subfolders(control_wb, base_dir: Path):
    ws = control_wb[SHEET_SCRIPT_MOVE]
    clear_table(ws, 2, 1, 1)
    row = 2
    for sub in sorted(base_dir.iterdir()):
        if sub.is_dir():
            ws.cell(row=row, column=1).value = ensure_trailing_slash(str(sub.resolve()))
            row += 1


def is_valid_cell_address(address: str) -> bool:
    return bool(re.fullmatch(r"[A-Za-z]{1,3}[0-9]{1,7}", address.strip()))


def run_batch_task(control_path: Path, base_override: Optional[Path], operation: str, append: bool = False):
    wb = ensure_control_workbook(control_path)
    ws_files = wb[SHEET_FILES]

    if base_override:
        ws_files[BASE_PATH_CELL].value = ensure_trailing_slash(str(base_override.resolve()))

    base_dir = get_control_base_dir(ws_files)
    selected_files = collect_selected_files(wb, base_dir)
    if not selected_files:
        wb.save(control_path)
        wb.close()
        print("선택된 파일이 없습니다. Include? 열 또는 필터(B4/B5)를 확인하세요.")
        return

    clear_status_columns(ws_files)
    wb.save(control_path)

    total = len(selected_files)
    start = time.time()
    log_messages: List[str] = []

    if operation == "value_find" and not append:
        clear_table(wb[SHEET_DATA_UPDATE], 2, 1, 6)

    for idx, entry in enumerate(selected_files, start=1):
        eta_text = ""
        if idx > 1:
            elapsed = time.time() - start
            avg = elapsed / (idx - 1)
            remaining = avg * (total - idx + 1)
            eta_text = f"(ETA {remaining/60:.1f} min)" if remaining >= 60 else f"(ETA {remaining:.0f} sec)"
        print(f"[{idx}/{total}] {entry.file_name} {eta_text}".strip())

        try:
            entry.processed_path = prepare_processed_copy(entry.original_path)
        except Exception as exc:
            mark_status(ws_files, entry.row, False, str(exc))
            wb.save(control_path)
            wb.close()
            log_path = write_log_if_needed([f"{entry.original_path} | {exc}"]) or ""
            raise SystemExit(f"백업 생성 중 오류: {exc}\n로그: {log_path}")

        try:
            if operation == "update_ids":
                message = task_update_ids(entry.processed_path)
            elif operation == "io_change":
                replacements = []
                ws_map = wb[SHEET_IO_NAME]
                for row in ws_map.iter_rows(min_row=1, max_col=2):
                    old = normalize_path(row[0].value)
                    new = row[1].value if len(row) > 1 else None
                    if old:
                        replacements.append((old, "" if new is None else str(new)))
                message = task_io_change(entry.processed_path, replacements)
            elif operation == "value_find":
                find_text = normalize_path(ws_files[FIND_VALUE_CELL].value)
                target_sheet = normalize_path(ws_files[TARGET_SHEET_CELL].value)
                if not find_text:
                    raise ValueError("B10 셀(찾을 문자열)이 비어 있습니다.")
                if not target_sheet:
                    raise ValueError("B12 셀(대상 시트)이 비어 있습니다.")
                message = task_value_find(wb, entry.processed_path, target_sheet, find_text, wb[SHEET_DATA_UPDATE])
            else:
                raise ValueError(f"알 수 없는 작업: {operation}")

            mark_status(ws_files, entry.row, True, message)
        except Exception as exc:
            mark_status(ws_files, entry.row, False, str(exc))
            log_messages.append(f"{entry.original_path} | {exc}")
            wb.save(control_path)
            wb.close()
            log_path = write_log_if_needed(log_messages)
            raise SystemExit(f"작업 중단: {exc}\n로그: {log_path}")

    wb.save(control_path)
    wb.close()
    log_path = write_log_if_needed(log_messages)
    elapsed = time.time() - start
    summary = f"{operation} 완료 ({total}개, {elapsed:.1f}초)"
    if log_path:
        summary += f"\n로그 파일: {log_path}"
    print(summary)


def cmd_scan(args):
    control = Path(args.control)
    base = Path(args.base).expanduser().resolve() if args.base else None
    count = list_excel_files(control, base)
    print(f"{count}개의 Excel 파일을 '{control}'에 기록했습니다.")


def cmd_update_ids(args):
    run_batch_task(Path(args.control), Path(args.base).resolve() if args.base else None, "update_ids")


def cmd_io_change(args):
    run_batch_task(Path(args.control), Path(args.base).resolve() if args.base else None, "io_change")


def cmd_value_find(args):
    run_batch_task(Path(args.control), Path(args.base).resolve() if args.base else None, "value_find", append=args.append)


def cmd_change_value(args):
    control_path = Path(args.control)
    wb = ensure_control_workbook(control_path)
    log_messages: List[str] = []
    success, failure = task_change_value(wb, log_messages)
    wb.save(control_path)
    wb.close()
    log_path = write_log_if_needed(log_messages)
    print(f"값 일괄 변경 완료: 성공 {success}건 / 실패 {failure}건")
    if log_path:
        print(f"오류 로그: {log_path}")


def cmd_list_subfolders(args):
    control_path = Path(args.control)
    wb = ensure_control_workbook(control_path)
    ws_files = wb[SHEET_FILES]
    if args.base:
        ws_files[BASE_PATH_CELL].value = ensure_trailing_slash(str(Path(args.base).resolve()))
    base_dir = get_control_base_dir(ws_files)
    list_subfolders(wb, base_dir)
    wb.save(control_path)
    wb.close()
    print(f"{base_dir} 하위 폴더 목록을 script_move 시트에 기록했습니다.")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="SMB Precheck 자동화 도구 (Python)")
    parser.add_argument("--control", default="control.xlsx", help="관리 통합문서 경로 (기본값: control.xlsx)")

    subparsers = parser.add_subparsers(dest="command", required=True)

    scan_parser = subparsers.add_parser("scan", help="폴더의 Excel 파일 목록을 갱신")
    scan_parser.add_argument("--base", help="기본 경로를 지정 (B2 셀 덮어쓰기)")
    scan_parser.set_defaults(func=cmd_scan)

    update_parser = subparsers.add_parser("update-ids", help="Test Case ID 재번호 작업")
    update_parser.add_argument("--base", help="기본 경로를 지정 (선택)")
    update_parser.set_defaults(func=cmd_update_ids)

    io_parser = subparsers.add_parser("io-change", help="IO 텍스트 치환 작업")
    io_parser.add_argument("--base", help="기본 경로를 지정 (선택)")
    io_parser.set_defaults(func=cmd_io_change)

    find_parser = subparsers.add_parser("value-find", help="값 찾기 (data_update 시트 갱신)")
    find_parser.add_argument("--base", help="기본 경로를 지정 (선택)")
    find_parser.add_argument("--append", action="store_true", help="기존 data_update 내용에 덧붙이기")
    find_parser.set_defaults(func=cmd_value_find)

    change_parser = subparsers.add_parser("change-value", help="data_update 기준으로 값 일괄 변경")
    change_parser.set_defaults(func=cmd_change_value)

    folder_parser = subparsers.add_parser("list-subfolders", help="하위 폴더를 script_move 시트에 기록")
    folder_parser.add_argument("--base", help="기본 경로를 지정 (선택)")
    folder_parser.set_defaults(func=cmd_list_subfolders)

    return parser


def main(argv: Optional[Sequence[str]] = None):
    parser = build_parser()
    args = parser.parse_args(argv)
    args.func(args)


if __name__ == "__main__":
    main()
